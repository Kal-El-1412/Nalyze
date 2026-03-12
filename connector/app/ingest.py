import csv
import logging
import os
import tempfile
from typing import Dict, Any, Optional
import duckdb

logger = logging.getLogger(__name__)


def _load_excel_to_duckdb(conn: duckdb.DuckDBPyConnection, file_path: str, ext: str):
    """Load an Excel file (.xlsx or .xls) into a DuckDB 'data' table with type inference."""
    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
    else:  # .xls
        import xlrd
        wb_xls = xlrd.open_workbook(file_path)
        ws_xls = wb_xls.sheet_by_index(0)
        datemode = wb_xls.datemode

        def _xls_rows():
            for row_idx in range(ws_xls.nrows):
                row = []
                for cell in ws_xls.row(row_idx):
                    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        row.append(None)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate.xldate_as_datetime(cell.value, datemode)
                            row.append(str(dt))
                        except Exception:
                            row.append(str(cell.value))
                    elif cell.ctype == xlrd.XL_CELL_NUMBER and cell.value == int(cell.value):
                        row.append(str(int(cell.value)))
                    else:
                        row.append(cell.value)
                yield tuple(row)

        rows_iter = _xls_rows()
        wb = wb_xls

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode='w', delete=False, suffix='.csv', newline=''
        ) as tmp_file:
            writer = csv.writer(tmp_file)
            for row in rows_iter:
                writer.writerow(row)
            tmp_path = tmp_file.name

        if ext == ".xlsx":
            wb.close()

        safe_tmp = tmp_path.replace("'", "''")
        conn.execute(f"""
            CREATE TABLE data AS
            SELECT * FROM read_csv_auto('{safe_tmp}',
                header=true,
                auto_detect=true,
                sample_size=-1
            )
        """)
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


class DataIngestor:
    def __init__(self):
        self.supported_extensions = {".csv", ".xlsx", ".xls", ".parquet"}

    def validate_file(self, file_path: str) -> bool:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        if not os.path.isfile(file_path):
            raise ValueError(f"Path is not a file: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in self.supported_extensions:
            raise ValueError(
                f"Unsupported file type: {ext}. Supported: {', '.join(self.supported_extensions)}"
            )

        return True

    async def analyze_file(self, file_path: str) -> Dict[str, Any]:
        self.validate_file(file_path)

        ext = os.path.splitext(file_path)[1].lower()
        logger.info(f"Analyzing file: {file_path}")

        conn = duckdb.connect(":memory:")

        try:
            safe_path = file_path.replace("'", "''")
            if ext == ".csv":
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{safe_path}')")
            elif ext in {".xlsx", ".xls"}:
                _load_excel_to_duckdb(conn, file_path, ext)
            elif ext == ".parquet":
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_parquet('{safe_path}')")

            result = conn.execute("SELECT COUNT(*) as row_count FROM data").fetchone()
            row_count = result[0] if result else 0

            columns_result = conn.execute("PRAGMA table_info('data')").fetchall()
            columns = [col[1] for col in columns_result]
            column_count = len(columns)

            logger.info(f"File analyzed: {row_count} rows, {column_count} columns")

            return {
                "row_count": row_count,
                "column_count": column_count,
                "columns": columns,
                "file_size": os.path.getsize(file_path)
            }

        except Exception as e:
            logger.error(f"Error analyzing file {file_path}: {e}")
            raise
        finally:
            conn.close()

    async def load_dataset(self, dataset_id: str, file_path: str) -> duckdb.DuckDBPyConnection:
        self.validate_file(file_path)

        ext = os.path.splitext(file_path)[1].lower()
        logger.info(f"Loading dataset {dataset_id} from {file_path}")

        conn = duckdb.connect(":memory:")

        try:
            safe_path = file_path.replace("'", "''")
            if ext == ".csv":
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{safe_path}')")
            elif ext in {".xlsx", ".xls"}:
                _load_excel_to_duckdb(conn, file_path, ext)
            elif ext == ".parquet":
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_parquet('{safe_path}')")

            logger.info(f"Dataset {dataset_id} loaded successfully")
            return conn

        except Exception as e:
            logger.error(f"Error loading dataset {dataset_id}: {e}")
            conn.close()
            raise


ingestor = DataIngestor()
