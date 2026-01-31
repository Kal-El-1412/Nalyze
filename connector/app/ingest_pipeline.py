import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import duckdb
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.storage import storage

logger = logging.getLogger(__name__)


class IngestionPipeline:
    def __init__(self):
        self.base_dir = Path.home() / ".cloaksheets" / "datasets"
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.max_xlsx_size_mb = 200
        self.chunk_size = 10000

    def get_dataset_dir(self, dataset_id: str) -> Path:
        dataset_dir = self.base_dir / dataset_id
        dataset_dir.mkdir(parents=True, exist_ok=True)
        return dataset_dir

    def get_db_path(self, dataset_id: str) -> Path:
        return self.get_dataset_dir(dataset_id) / "db.duckdb"

    def get_catalog_path(self, dataset_id: str) -> Path:
        return self.get_dataset_dir(dataset_id) / "catalog.json"

    def _get_file_size_mb(self, file_path: str) -> float:
        return os.path.getsize(file_path) / (1024 * 1024)

    def _get_file_extension(self, file_path: str) -> str:
        return Path(file_path).suffix.lower()

    async def ingest(self, dataset_id: str, file_path: str, job_id: str, force: bool = False):
        ext = self._get_file_extension(file_path)

        if ext == ".csv":
            await self.ingest_csv(dataset_id, file_path, job_id)
        elif ext in [".xlsx", ".xls"]:
            await self.ingest_xlsx(dataset_id, file_path, job_id, force)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    async def ingest_csv(self, dataset_id: str, file_path: str, job_id: str):
        logger.info(f"Starting ingestion for dataset {dataset_id} from {file_path}")

        try:
            await storage.update_job(
                job_id=job_id,
                status="running",
                started_at=datetime.utcnow().isoformat()
            )

            db_path = self.get_db_path(dataset_id)
            catalog_path = self.get_catalog_path(dataset_id)

            if db_path.exists():
                db_path.unlink()
                logger.info(f"Removed existing database at {db_path}")

            conn = duckdb.connect(str(db_path))

            logger.info(f"Loading CSV from {file_path} into DuckDB")
            conn.execute(f"""
                CREATE TABLE data AS
                SELECT * FROM read_csv_auto('{file_path}',
                    sample_size=-1,
                    ignore_errors=false,
                    auto_detect=true
                )
            """)

            logger.info("CSV loaded successfully, generating catalog")
            catalog = self._generate_catalog(conn)

            with open(catalog_path, 'w') as f:
                json.dump(catalog, f, indent=2)

            conn.close()
            logger.info(f"Catalog saved to {catalog_path}")

            await storage.update_dataset(
                dataset_id=dataset_id,
                updates={
                    "status": "ingested",
                    "lastIngestedAt": datetime.utcnow().isoformat()
                }
            )

            await storage.update_job(
                job_id=job_id,
                status="done",
                finished_at=datetime.utcnow().isoformat()
            )

            logger.info(f"Ingestion completed successfully for dataset {dataset_id}")

        except Exception as e:
            logger.error(f"Ingestion failed for dataset {dataset_id}: {e}", exc_info=True)

            await storage.update_dataset(
                dataset_id=dataset_id,
                updates={"status": "error"}
            )

            await storage.update_job(
                job_id=job_id,
                status="error",
                finished_at=datetime.utcnow().isoformat(),
                error=str(e)
            )

    async def ingest_xlsx(self, dataset_id: str, file_path: str, job_id: str, force: bool = False):
        logger.info(f"Starting XLSX ingestion for dataset {dataset_id} from {file_path}")

        try:
            file_size_mb = self._get_file_size_mb(file_path)
            logger.info(f"XLSX file size: {file_size_mb:.2f} MB")

            if file_size_mb > self.max_xlsx_size_mb and not force:
                error_msg = (
                    f"XLSX file is {file_size_mb:.2f} MB, which exceeds the recommended limit of "
                    f"{self.max_xlsx_size_mb} MB. For better performance, please export to CSV format. "
                    f"Alternatively, you can force ingestion by adding ?force=true to the request."
                )
                raise ValueError(error_msg)

            await storage.update_job(
                job_id=job_id,
                status="running",
                started_at=datetime.utcnow().isoformat()
            )

            db_path = self.get_db_path(dataset_id)
            catalog_path = self.get_catalog_path(dataset_id)

            if db_path.exists():
                db_path.unlink()
                logger.info(f"Removed existing database at {db_path}")

            conn = duckdb.connect(str(db_path))

            logger.info(f"Loading XLSX from {file_path}")
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)

            sheet = self._select_best_sheet(wb)
            logger.info(f"Selected sheet: {sheet.title}")

            self._ingest_sheet_to_duckdb(conn, sheet)
            wb.close()

            logger.info("XLSX loaded successfully, generating catalog")
            catalog = self._generate_catalog(conn)

            with open(catalog_path, 'w') as f:
                json.dump(catalog, f, indent=2)

            conn.close()
            logger.info(f"Catalog saved to {catalog_path}")

            await storage.update_dataset(
                dataset_id=dataset_id,
                updates={
                    "status": "ingested",
                    "lastIngestedAt": datetime.utcnow().isoformat()
                }
            )

            await storage.update_job(
                job_id=job_id,
                status="done",
                finished_at=datetime.utcnow().isoformat()
            )

            logger.info(f"XLSX ingestion completed successfully for dataset {dataset_id}")

        except Exception as e:
            logger.error(f"XLSX ingestion failed for dataset {dataset_id}: {e}", exc_info=True)

            await storage.update_dataset(
                dataset_id=dataset_id,
                updates={"status": "error"}
            )

            await storage.update_job(
                job_id=job_id,
                status="error",
                finished_at=datetime.utcnow().isoformat(),
                error=str(e)
            )

    def _select_best_sheet(self, workbook):
        if not workbook.sheetnames:
            raise ValueError("Workbook contains no sheets")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if hasattr(sheet, 'tables') and sheet.tables:
                logger.info(f"Found named table in sheet: {sheet_name}")
                return sheet

        first_sheet = workbook[workbook.sheetnames[0]]
        return first_sheet

    def _ingest_sheet_to_duckdb(self, conn: duckdb.DuckDBPyConnection, sheet):
        rows_iter = sheet.iter_rows(values_only=True)

        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError("Sheet is empty or has no data")

        headers = []
        for idx, cell in enumerate(header_row):
            if cell is None or str(cell).strip() == "":
                headers.append(f"column_{idx + 1}")
            else:
                headers.append(str(cell).strip())

        logger.info(f"Detected {len(headers)} columns: {headers[:5]}...")

        conn.execute("CREATE TABLE data (" + ", ".join([f'"{h}" VARCHAR' for h in headers]) + ")")

        chunk = []
        row_count = 0

        for row in rows_iter:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            cleaned_row = []
            for cell in row:
                if cell is None:
                    cleaned_row.append(None)
                else:
                    cleaned_row.append(str(cell))

            chunk.append(cleaned_row)
            row_count += 1

            if len(chunk) >= self.chunk_size:
                self._insert_chunk(conn, headers, chunk)
                logger.info(f"Inserted chunk of {len(chunk)} rows (total: {row_count})")
                chunk = []

        if chunk:
            self._insert_chunk(conn, headers, chunk)
            logger.info(f"Inserted final chunk of {len(chunk)} rows (total: {row_count})")

        logger.info(f"Total rows inserted: {row_count}")

    def _insert_chunk(self, conn: duckdb.DuckDBPyConnection, headers: List[str], chunk: List[List]):
        placeholders = ", ".join(["?" for _ in headers])
        col_names = ", ".join([f'"{h}"' for h in headers])
        insert_sql = f"INSERT INTO data ({col_names}) VALUES ({placeholders})"

        conn.executemany(insert_sql, chunk)

    def _generate_catalog(self, conn: duckdb.DuckDBPyConnection) -> Dict[str, Any]:
        row_count = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]

        columns_info = conn.execute("""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_name = 'data'
            ORDER BY ordinal_position
        """).fetchall()

        columns = []
        basic_stats = {}
        detected_date_columns = []
        detected_numeric_columns = []

        for col_name, col_type in columns_info:
            columns.append({"name": col_name, "type": col_type})

            col_type_lower = col_type.lower()

            if any(t in col_type_lower for t in ['int', 'double', 'float', 'decimal', 'numeric', 'real']):
                detected_numeric_columns.append(col_name)
                stats = self._get_numeric_stats(conn, col_name, row_count)
                basic_stats[col_name] = stats

            elif any(t in col_type_lower for t in ['date', 'timestamp', 'time']):
                detected_date_columns.append(col_name)
                stats = self._get_date_stats(conn, col_name, row_count)
                basic_stats[col_name] = stats

            else:
                stats = self._get_text_stats(conn, col_name, row_count)
                basic_stats[col_name] = stats

        catalog = {
            "table": "data",
            "rowCount": row_count,
            "columns": columns,
            "basicStats": basic_stats,
            "detectedDateColumns": detected_date_columns,
            "detectedNumericColumns": detected_numeric_columns
        }

        return catalog

    def _get_numeric_stats(self, conn: duckdb.DuckDBPyConnection, col_name: str, total_rows: int) -> Dict[str, Any]:
        try:
            result = conn.execute(f"""
                SELECT
                    MIN("{col_name}") as min_val,
                    MAX("{col_name}") as max_val,
                    AVG("{col_name}") as avg_val,
                    COUNT(*) FILTER (WHERE "{col_name}" IS NULL) as null_count
                FROM data
            """).fetchone()

            null_count = result[3] or 0
            null_pct = (null_count / total_rows * 100) if total_rows > 0 else 0

            return {
                "min": result[0],
                "max": result[1],
                "avg": result[2],
                "nullPct": round(null_pct, 2)
            }
        except Exception as e:
            logger.warning(f"Error getting numeric stats for {col_name}: {e}")
            return {"nullPct": 0}

    def _get_date_stats(self, conn: duckdb.DuckDBPyConnection, col_name: str, total_rows: int) -> Dict[str, Any]:
        try:
            result = conn.execute(f"""
                SELECT
                    MIN("{col_name}") as min_val,
                    MAX("{col_name}") as max_val,
                    COUNT(*) FILTER (WHERE "{col_name}" IS NULL) as null_count
                FROM data
            """).fetchone()

            null_count = result[2] or 0
            null_pct = (null_count / total_rows * 100) if total_rows > 0 else 0

            return {
                "min": str(result[0]) if result[0] else None,
                "max": str(result[1]) if result[1] else None,
                "nullPct": round(null_pct, 2)
            }
        except Exception as e:
            logger.warning(f"Error getting date stats for {col_name}: {e}")
            return {"nullPct": 0}

    def _get_text_stats(self, conn: duckdb.DuckDBPyConnection, col_name: str, total_rows: int) -> Dict[str, Any]:
        try:
            result = conn.execute(f"""
                SELECT
                    COUNT(*) FILTER (WHERE "{col_name}" IS NULL) as null_count,
                    APPROX_COUNT_DISTINCT("{col_name}") as distinct_count
                FROM data
            """).fetchone()

            null_count = result[0] or 0
            null_pct = (null_count / total_rows * 100) if total_rows > 0 else 0

            return {
                "nullPct": round(null_pct, 2),
                "approxDistinct": result[1]
            }
        except Exception as e:
            logger.warning(f"Error getting text stats for {col_name}: {e}")
            return {"nullPct": 0, "approxDistinct": 0}

    async def load_catalog(self, dataset_id: str) -> Dict[str, Any]:
        catalog_path = self.get_catalog_path(dataset_id)

        if not catalog_path.exists():
            raise FileNotFoundError(f"Catalog not found for dataset {dataset_id}")

        with open(catalog_path, 'r') as f:
            return json.load(f)


ingestion_pipeline = IngestionPipeline()
